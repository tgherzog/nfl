#!/usr/local/bin/python

"""
NFL stats: a program to calculate NFL team standings

Script is now designed as a module. To get started, use the command line options or load
interactively and try:

from nfl import NFL
nfl = NFL()
nfl.load('NFLData.xlsx')

nfl('MIN')                       # info and standings for vikings
nfl('NFC')                       # or the NFC conference
nfl.wlt(['MIN', 'GB'])           # win/loss/tie info
nfl.tiebreakers(['DAL', 'PHI'])  # tiebreaker calculations

# and lots more, and yet lots of room for enhancements/improvements ;)

Usage:
    nfl.py update FILE WEEK
    nfl.py team TEAM [--file FILE]
    nfl.py tiebreakers TEAMS... [--file FILE]

Commands:
    update          Update the database FILE, scraping data up to and including WEEK

    team            Report TEAM stats and schedule (division or conference work too)

    tiebreakers     Report tiebreaker analysis for the specified TEAMS

Options:
    --file FILE     Use FILE as database path [default: NFLData.xlsx]

"""


import openpyxl
import pprint
import requests
import sys
import re
import logging
import time
from docopt import docopt
from pyquery import PyQuery
import urllib
from datetime import datetime, timedelta
from dateutil import tz
import numpy as np
import pandas as pd

class NFLTeam():
    def __init__(self, code, host):

        elem = host.teams_[code]
        self.code = code
        self.name = elem['name']
        self.div  = elem['div']
        self.conf = elem['conf']
        self.host = host

    @property
    def division(self):
        '''Return set of teams in this team's division
        '''
        return self.host.divs_[self.div]

    @property
    def conference(self):
        '''Return set of teams in this team's conference
        '''
        return self.host.confs_[self.conf]

    @property
    def standings(self):
        '''Return team standings
        '''

        return self.host.team_stats(self.code)

    @property
    def schedule(self):
        '''Return the team's schedule
        '''

        return self.host.schedule(self.code, by='team')

    @property
    def opponents(self):
        '''The team's opponents
        '''

        return self.host.opponents(self.code)

    def boxscore(self, week=None):

        if week is None:
            week = self.host.week

        return self.host.engine.boxscore(self.host, self.code, week)

    def __repr__(self):
        wk = self.host.week or 1
        s  = self.schedule.loc[range(1,wk+3)].__repr__()
        return '{}: {} ({})\n'.format(self.code, self.name, self.div) + self.standings.__repr__() + '\n' + s

    def _repr_html_(self):
        wk = self.host.week or 1
        s  = self.schedule.loc[range(1,wk+3)]._repr_html_()
        return '<h3>{}: {} ({})</h3>\n'.format(self.code, self.name, self.div) + self.standings._repr_html_() + '\n' + s


class NFLDivision():
    def __init__(self, code, host):

        self.code = code
        self.host = host
        self.teams = host.divs_[code]

    @property
    def standings(self):
        ''' Return division standings

            rank: if True, add 'rank' column based on in-division tiebreaker rules (longer)
                  if False, sort results by game wlt
        '''

        z = self.host.standings
        z = z[z['div']==self.code]
        return z

    def __repr__(self):
        # ensures tiebreakers are used to determine order, then deletes the rank column
        return '{}\n'.format(self.code) + self.standings.__repr__()

    def _repr_html_(self):
        return '<h3>{}</h3>\n'.format(self.code) + self.standings._repr_html_()

class NFLConference():
    def __init__(self, code, host):

        self.code = code
        self.host = host
        self.teams = host.confs_[code]

    @property
    def divisions(self):
        return set(map(lambda x: '-'.join([self.code, x]), ['North', 'East', 'South', 'West']))


    @property
    def standings(self):

        z = self.host.standings
        return z[z['div'].str.startswith(self.code)]
        return z


    def playoffs(self, count=7):
        '''Return the current conference seeds in order
        '''

        # get top-listed teams from each division
        champs = self.standings.reset_index().groupby(('div','')).first()['team']

        # restructure and reorder by seed
        champs = pd.Series(map(lambda x: x.split('-')[-1], champs.index), index=champs, name=self.code)
        champs = champs.reindex(self.host.tiebreaks(champs.index).index)

        # pull wildcards from the remaining teams
        wc = self.host.tiebreaks(self.teams - set(champs.index))
        for elem in wc.index[:count-len(champs)]:
            champs[elem] = 'Wildcard'

        return champs



    def __repr__(self):
        return '{}\n'.format(self.code) + self.standings.__repr__()

    def _repr_html_(self):
        return '<h3>{}</h3>\n'.format(self.code) + self.standings._repr_html_()

class NFL():
    '''Impelementation class for NFL data operations

       By default calls to scoreboard update the game database in real-time.
       Set autoUpdate=False to disable
    '''

    year = None
    week = None
    path = None
    autoUpdate = True

    def __init__(self, year=None, engine=None):
        self.teams_  = {}
        self.iteams_ = {}  # team name to id
        self.divs_   = {}
        self.confs_  = {}
        self.games_  = []
        self.max_week = 0
        self.stats = pd.DataFrame()
        self.year = year
        self.engine = engine
        if not self.year:
            dt = datetime.now()
            self.year = dt.year if dt.month >= 4 else dt.year-1

        if not engine:
            self.engine = NFLSourceESPN()
        elif type(engine) is str:
            self.engine = getattr(sys.modules[__name__], engine)()

    def __call__(self, i):

        if i in self.confs_:
            return NFLConference(i, self)

        if i in self.divs_:
            return NFLDivision(i, self)

        if i in self.teams_:
            return NFLTeam(i, self)

    def __repr__(self):

        return self.standings.__repr__()

    def _repr_html_(self):

        return self.standings._repr_html()

    def load(self, path=None):
        ''' Loads data from the specified excel file

            path:   path to Excel file
        '''

        if not path:
            path = 'NFLData-{}.xlsx'.format(self.year)

        # in case of reload
        self.path   = path
        self.games_ = []
        self.max_week = 0                  

        wb = openpyxl.load_workbook(path, read_only=True)
        self.load_league(wb)

        for row in wb['Scores']:
            if row[0].row > 1 and row[0].value:
                # at/ht = away team/home team - same for scores
                game = {'wk': row[0].value,
                    'at': row[1].value,
                    'as': row[2].value if row[2].value is not None else np.nan,
                    'ht': row[3].value,
                    'hs': row[4].value if row[4].value is not None else np.nan,
                    'ts': row[5].value,
                    'id': row[6].value,
                }
                game['p'] = game['as'] is not np.nan and game['hs'] is not np.nan
                self.games_.append(game)
                self.max_week = max(self.max_week, game['wk'])

        # sort games chronologically
        self.games_ = sorted(self.games_, key=lambda x: x['ts'])

        # reset the engine
        my_engine   = self.engine.__class__.__name__
        file_engine = wb['Meta'].cell(row=2, column=2).value
        if file_engine and my_engine != file_engine:
            engine_class = getattr(sys.modules[__name__], file_engine)
            self.engine = engine_class()

        file_year = wb['Meta'].cell(row=3, column=2).value
        if not file_year:
            self.year = file_year

        if not self.week:
            # assumes games are sorted chronologically, per above
            now = datetime.now().replace(hour=0, minute=0, second=0)
            weekends = {}
            for row in self.games_:
                t = row['ts'].replace(hour=0, minute=0, second=0)
                weekends[row['wk']] = max(weekends.get(row['wk'], datetime.min), t)
                if t <= now:
                    self.week = row['wk']

            if self.week and self.week < self.max_week and now > weekends[self.week]:
                # if between weeks, go to the next week
                self.week += 1

        self.stats = None
        return self

    def load_league(self, wb):
        '''Loads teams, conferences and divisions from the supplied spreadsheet. Called by load() and build()
        '''

        self.teams_ = {}
        self.divs_  = {}
        self.confs_ = {}

        for row in wb['Divisions']:
            team = row[0].value
            conf = row[2].value
            div = row[3].value
            alt = row[4].value
            if team and conf and div:
                div = '-'.join([conf, div])
                self.teams_[team] = {'name': row[1].value, 'div': div, 'conf': conf, 'alt': alt}
                if self.divs_.get(div) is None:
                    self.divs_[div] = set()

                self.divs_[div].add(team)

                if self.confs_.get(conf) is None:
                    self.confs_[conf] = set()

                self.confs_[conf].add(team)


        self.iteams_ = {v['name']:k for k,v in self.teams_.items()}

    def teamid(self, name):
        return self.iteams_.get(name)

    def _stats(self):
        '''Return the master statistics table, building it first if necessary. Teams are ordered by conference,
        division and division rank. Changes to raw game scores (via load, set, etc) invalidate the table
        so the next call to _stats rebuilds it
        '''

        if self.stats is not None:
            return self.stats

        stat_cols = [('name',''),('div',''),('conf','')]
        stat_cols += list(pd.MultiIndex.from_product([['overall','division','conference', 'vic_stren', 'sch_stren'],['win','loss','tie','pct']]))
        stat_cols += list(pd.MultiIndex.from_product([['misc'],['rank-conf','rank-overall', 'pts-scored', 'pts-allowed', 'conf-pts-scored', 'conf-pts-allowed']]))
        stats = pd.DataFrame(columns=pd.MultiIndex.from_tuples(stat_cols))
        stats.index.name = 'team'

        for k,team in self.teams_.items():
            stats.loc[k, ['name', 'div', 'conf']] = (team['name'], team['div'], team['conf'])
            stats.loc[k, ['overall','division','conference', 'vic_stren', 'sch_stren', 'misc']] = 0

        # special dataframe of game info to streamline processing. Contains 2 rows per game for
        # outcomes from each team's perspective. Perhaps someday this would be useful elsewhere
        games = pd.DataFrame(index=range(len(self.games_)*2), columns=['week','team','opp','wlt', 'scored', 'allowed'])
        z = 0
        for game in self.games_:
            if game['p']:
                games.loc[z]   = [game['wk'], game['ht'], game['at'], NFL.result(game['hs'], game['as']), game['hs'], game['as']]
                games.loc[z+1] = [game['wk'], game['at'], game['ht'], NFL.result(game['as'], game['hs']), game['as'], game['hs']]
                z += 2

        games.dropna(inplace=True)

        for k,row in games.iterrows():
            stats.loc[row.team][('overall',row.wlt)] += 1
            stats.loc[row.team][('misc','pts-scored')] += row.scored
            stats.loc[row.team][('misc','pts-allowed')] += row.allowed
            if self.teams_[row.team]['div'] == self.teams_[row.opp]['div']:
                stats.loc[row.team][('division',row.wlt)] += 1

            if self.teams_[row.team]['conf'] == self.teams_[row.opp]['conf']:
                stats.loc[row.team][('conference',row.wlt)] += 1
                stats.loc[row.team][('misc','conf-pts-scored')] += row.scored
                stats.loc[row.team][('misc','conf-pts-allowed')] += row.allowed

        # strength of victory/schedule
        for (k,row) in stats.iterrows():
            # calculate strength of schedule and copy
            t = games[games.team==k]['opp']
            stats.loc[k]['sch_stren'] = stats.loc[t]['overall'].sum()

            # same for strength of victory
            t = games[(games.team==k) & (games.wlt=='win')]['opp']
            stats.loc[k]['vic_stren'] = stats.loc[t]['overall'].sum()

        # temporary table for calculating ranks - easier syntax
        t = pd.concat([stats['conf'], stats['misc'][['pts-scored','pts-allowed']]], axis=1)
        stats[('misc','rank-overall')] = (t['pts-scored'].rank() + t['pts-allowed'].rank(ascending=False)).rank()

        t['conf-off-rank'] = t.groupby('conf')['pts-scored'].rank()
        t['conf-def-rank'] = t.groupby('conf')['pts-allowed'].rank(ascending=False)
        t['conf-rank'] = t['conf-off-rank'] + t['conf-def-rank']
        stats[('misc', 'rank-conf')] = t.groupby('conf')['conf-rank'].rank()

        for i in ['overall','division','conference', 'sch_stren', 'vic_stren']:
            stats[(i,'pct')] = (stats[(i,'win')] + stats[(i,'tie')]*0.5) / stats[i].sum(axis=1).replace({0: np.nan})

        # sort by division tiebreakers
        self.stats = stats           # so that tiebreaks doesn't go recursive

        s = pd.Series(index=stats.index)
        for div in stats['div'].unique():
            z = self.tiebreaks(div)
            z[:] = range(len(z))
            s.loc[z.index] = z.astype(s.dtype)

        self.stats = stats.assign(divrank=s).sort_values(['div','divrank']).drop('divrank', level=0, axis=1)
        return self.stats

    def reload(self):
        ''' Reloads the previous Excel file
        '''

        if self.path:
            self.load(self.path)

        return self

    def update(self, path=None):
        ''' Updates the Scores sheet of the specified Excel file by scraping the URL shown in code.
            The file must already exist and have a valid "Divisions" sheet with team names consistent
            with those on the source website.  All but the first row of the Scores sheet will be replaced.

            path:   path to Excel workbook

            year:   season to load; else infer the latest season from the current date
        '''

        if not path:
            path = 'NFLData-{}.xlsx'.format(self.year)

        wb = openpyxl.load_workbook(path, read_only=False)
        self.load_league(wb)
        tids = {v['name']:k for k,v in self.teams_.items()}
        ws = wb['Scores']
        ws.delete_rows(2, ws.max_row)
        row = 2

        for elem in self.engine.games(self):
            hteam = self.teamid(elem['hteam'])
            ateam = self.teamid(elem['ateam'])
            if not hteam:
                raise ValueError('"{}" is not a recognized team name'.format(elem['hteam']))

            if not ateam:
                raise ValueError('"{}" is not a recognized team name'.format(elem['ateam']))

            ws.cell(row=row, column=1, value=elem['week'])
            ws.cell(row=row, column=2, value=ateam)
            ws.cell(row=row, column=4, value=hteam)
            ws.cell(row=row, column=6, value=elem['date'])

            if type(elem['ascore']) is int:
                ws.cell(row=row, column=3, value=elem['ascore'])

            if type(elem['hscore']) is int:
                ws.cell(row=row, column=5, value=elem['hscore'])

            ws.cell(row=row, column=7, value=elem.get('id'))
            row += 1

        ws = wb['Meta']
        ws.cell(row=1, column=2, value=datetime.now())
        ws.cell(row=2, column=2, value=type(self.engine).__name__)
        ws.cell(row=3, column=2, value=self.year)

        wb.save(path)
        wb.close()
        self.load(path)
        return self

    @property
    def scoreboard(self):

        z = self.engine.scoreboard(self)
        if z and self.autoUpdate and self.year == z.year:
            for (k,v) in z.scoreboard.iterrows():
                if v['status'] == 'Final' or v['status'] == 'Final/OT':
                    self.set(z.week, **{v['hteam']: v['hscore'], v['ateam']: v['ascore']})

        return z

    @property
    def standings(self):

        return self._stats()[['name','div','overall','division','conference']]

    def set(self, wk, **kwargs):
        ''' Set the final score(s) for games in a given week. You can use this to create
            hypothetical outcomes and analyze the effect on team rankings. Scores
            are specified by team code and applied to the specified week's schedule.
            If the score is specified for only one team in a game, the score of the other
            team is assumed to be zero if not previously set.

            wk:         week number
            **kwargs    dict of team codes and final scores

            # typical example
            set(7, MIN=17, GB=10)

            The above will set the score for the MIN/GB game in week 7 if
            MIN and GB play each other in that week. Otherwise, it will
            set each team's score respectively and default their opponent's scores
            to 0. Scores for bye teams in that week are ignored.
        '''

        if type(wk) is dict:
            for k,v in wk.items():
                self.set(k, **v)

            return

        # sanity checks
        bogus = set(kwargs.keys()) - set(self.teams_.keys())
        if len(bogus) > 0:
            raise KeyError('Invalid team codes: {}'.format(','.join(bogus)))       


        for elem in self.games_:
            if wk == elem['wk']:
                if elem['ht'] in kwargs:
                    if kwargs[elem['ht']] < 0:
                        # tie with unspecified score
                        elem['hs'] = elem['as'] = 0
                    else:
                        elem['hs'] = kwargs[elem['ht']]
                        if elem['as'] is None:
                            elem['as'] = 0 if elem['hs'] > 0 else 1

                    elem['p'] = True

                if elem['at'] in kwargs:
                    if kwargs[elem['at']] < 0:
                        elem['hs'] = elem['as'] = 0
                    else:
                        elem['as'] = kwargs[elem['at']]
                        if elem['hs'] is None:
                            elem['hs'] = 0 if elem['as'] > 0 else 1

                    elem['p'] = True

            elif wk < elem['wk']:
                # assuming elements are sorted by week, we can stop at this point
                break

        self.stats = None # signal to rebuild stats

    def clear(self, week, teams=None):
        '''Clear scores for a given week or weeks

        week:   can be an integer, range or list-like. Pass None to clear all (for whatever reason)

        teams:  limit operation to games for specified teams (list-like)
        '''

        if type(week) is int:
            week = [week]

        for elem in self.games_:
            if week is None or elem['wk'] in week:
                if teams is None or elem['ht'] in teams or elem['at'] in teams:
                    elem['p'] = False
                    elem['hs'] = elem['as'] = None

        self.stats = None


    def games(self, teams=None, limit=None, allGames=False):
        ''' generator to iterate over score data

            teams:      code or list-like of teams to fetch

            limit:      range or list-like of weeks to fetch. Integers are converted
                        to the top limit of a range

            Example:

                for score in scores('MIN', limit=10) # fetch Vikings record up to but not including week 10
        '''

        if type(limit) is int:
            limit = range(limit, limit+1)

        teams = self._teams(teams)

        for elem in self.games_:
            if limit is None or elem['wk'] in limit:
                if teams is None or elem['at'] in teams or elem['ht'] in teams:
                    if elem['p'] or allGames:
                        yield elem


    def game(self, team, week):
        ''' return a single game for the given team and week 
        '''

        for elem in self.games(team, week, allGames=True):
            return elem


    def scores(self, teams=None, limit=None):
        ''' Returns interated game data structured by teams

            Result is  dict keyed by team code each of game results as follows:

            [wlt, us, them, op, home, week]

            wlt:  'win' 'loss' or 'tie'
            us:   our final score
            them: their final score
            op:   opponent (team code)
            home: True for home games
            week: week number
        '''

        if teams is None:
            z = {i:[] for i in self.teams_.keys()}
        else:
            teams = self._teams(teams)
            z = {i:[] for i in teams}

        for game in self.games(teams, limit):
            if teams is None or game['at'] in teams:
                z[game['at']].append([NFL.result(game['as'], game['hs']), game['as'], game['hs'], game['ht'], False, game['wk']])

            if teams is None or game['ht'] in teams:
                z[game['ht']].append([NFL.result(game['hs'], game['as']), game['hs'], game['as'], game['at'], True, game['wk']])

        return z


    def schedule(self, teams=None, weeks=None, by='game', ts=False):
        ''' Return schedule for a week or team. If a single week and
        team are requested the result is a Series. If multiple weeks/teams
        are requested the result is a DataFrame, Multiple weeks and teams
        together will return a multi-indexed DataFrame.

        teams:      teams to return: a single team code, division name, conference name or team list
                    Pass None for all teams

        weeks:      week to return or list-like of week numbers
                    Pass None for all weeks

        by:         Data orientation, either 'team' or 'game'
                        'game' will return unique games based on the team(s) you request.
                        If two of the teams you request play each other the result will
                        be one row not two

                        'team' will return teams as rows; if two requested teams play
                        each other there will be two rows not one. Additionally, 'bye'
                        teams will be listed as empty rows

        ts:         Return dates as Pandas datetime objects that can be used computationally,
                    otherwise (default) return an easily read string

        '''

        def to_date(obj):
            obj = pd.to_datetime(obj)
            if ts:
                return obj

            return '{}/{} {:02d}:{:02d}'.format(obj.month, obj.day, obj.hour, obj.minute)

        teams2 = self._teams(teams)
        if type(teams) is str and len(teams2) > 1:
            # This is so we differentiate between a division/conference name and a team code
            teams = teams2

        weeks2 = self._teams(weeks)

        if by == 'game':
            df = pd.DataFrame(columns=['week', 'hteam', 'ateam', 'hscore', 'ascore', 'date'])
            for game in self.games(teams=teams2, limit=weeks2, allGames=True):
                df.loc[len(df)] = [game['wk'], game['ht'], game['at'], game['hs'], game['as'], to_date(game['ts'])]

            if type(teams) is str and type(weeks) is int:
                # need a special test here to make sure we have a data frame
                if len(df) > 0:
                    return df.drop('week', axis=1).iloc[0]
                else:
                    # return an empty series
                    return pd.Series(index=df.columns.drop('week'))

            elif type(weeks) is int:
                return df.set_index('hteam').drop('week', axis=1)

            return df.set_index(['week', 'hteam'])

        if teams2 is None:
            teams2 = list(self.teams_.keys())
            teams2.sort()

        # Here we construct the database index in advance so that it includes empty
        # rows for teams with bye weeks
        if weeks2 is None:
            weeks2 = range(1, self.max_week+1)

        df = pd.DataFrame(index=pd.MultiIndex.from_product([weeks2, teams2], names=['week', 'team']), columns=['opp', 'loc', 'score', 'opp_score', 'wlt', 'date'])
        for game in self.games(teams=teams2, limit=weeks2, allGames=True):
            if game['ht'] in teams2:
                df.loc[(game['wk'],game['ht'])] = [game['at'], 'home', game['hs'], game['as'], NFL.result(game['hs'], game['as']), to_date(game['ts'])]
            
            if game['at'] in teams2:
                df.loc[(game['wk'],game['at'])] = [game['ht'], 'away', game['as'], game['hs'], NFL.result(game['as'], game['hs']), to_date(game['ts'])]

        if type(teams) is str and type(weeks) is int:
            return df.loc[(weeks,teams)]
        elif type(teams) is str:
            return df.xs(teams, level=1)
        elif type(weeks) is int:
            return df.xs(weeks, level=0)

        return df


    def opponents(self, teams, limit=None):
        ''' Returns the set of common opponents of one or more teams

            The teams argument can be a single team or a list.
        '''

        if teams is None:
            raise ValueError("teams cannot be None here")

        teams = self._teams(teams)

        ops = {t:set() for t in teams}

        for game in self.games(teams, limit):
            if game['ht'] in ops:
                ops[game['ht']].add(game['at'])

            if game['at'] in ops:
                ops[game['at']].add(game['ht'])

        # Resulting set is common (intersection) of opponents excluding the teams themselves
        z = None
        for s in ops.values():
            if z is None:
                z = s
            else:
                z &= s

        z -= set(teams)
        return z

    def wlt(self, teams=None, within=None, limit=None):
        '''Return the wlt stats of one or more teams

        teams:  team code or list of team codes

        within: list of team codes that defines the wlt universe
        '''

        return self._wlt(teams=teams, within=within, limit=limit)[0].drop(['scored','allowed'], axis=1)

    def matrix(self, teams=None, limit=None, allGames=False):
        '''Return a matrix of teams and the number of games played against each other
        '''

        return self._wlt(teams, limit=limit, allGames=allGames)[1]


    def _wlt(self, teams=None, within=None, limit=None, allGames=False):
        ''' Internal function for calculating wlt from games database with
        options to calculate ancillary data.

        points: include columns for points scored and allowed

        matrix: if True, returns the wlt frame and the games frame as a tuple
        '''

        teams = self._teams(teams)

        cols = ['win','loss','tie', 'pct', 'scored','allowed']

        df = pd.DataFrame(index=list(teams), columns=cols)
        df[df.columns] = 0
        df.columns.name = 'outcome'
        df.index.name = 'team'

        # define a matrix of games played against opponents
        m = pd.DataFrame(index=list(teams), columns=list(teams))
        m[m.columns] = 0
        for t in teams:
            m.loc[t, t] = np.nan

        for game in self.games(teams, limit, allGames):
            if game['ht'] in teams and (within is None or game['at'] in within):
                z = NFL.result(game['hs'], game['as'])
                if z:
                    df.loc[game['ht'], z] += 1
                    df.loc[game['ht'], 'scored'] += game['hs']
                    df.loc[game['ht'], 'allowed'] += game['as']

                if game['at'] in m.columns:
                    m.loc[game['ht'], game['at']] += 1

            if game['at'] in teams and (within is None or game['ht'] in within):
                z = NFL.result(game['as'], game['hs'])
                if z:
                    df.loc[game['at'], z] += 1
                    df.loc[game['at'], 'scored'] += game['as']
                    df.loc[game['at'], 'allowed'] += game['hs']


                if game['ht'] in m.columns:
                    m.loc[game['at'], game['ht']] += 1                


        df['pct'] = (df['win'] + df['tie'] * 0.5) / df.drop(columns=['scored','allowed'], errors='ignore').sum(axis=1)
        df.sort_values('pct', ascending=False, inplace=True)
        
        return (df, m)

    def team_stats(self, team):
        '''Return stats for a single team
        '''

        return self._stats().loc[team][['overall','division','conference']].unstack()


    def tiebreakers(self, teams):
        '''Return tiebreaker analysis for specified teams

        Each row in the returned dataframe is the results of a step in the NFL's tiebreaker procedure
        currently defined here: https://www.nfl.com/standings/tie-breaking-procedures

        Rows are in order of precedence and depend on whether the teams are in the same division or not

        rank (conference or overall) statistics are always in increasing order, e.g. 1 is the worst
        ranked team

        An occurrence of 'inf' as a value indicates that comparisons are not valid for that rule
        in the given context as per the tiebreakers procedures. The win/loss/tie statistics are still
        provided for information. For example, if the specified teams do not all have the same number
        of head-to-head matchups, or a wild-card tiebreaker does not have at least 4 common games, the
        'pct' column is set to 'inf' for those rules.

        Example:

        z = nfl.tiebreakers(nfl('NFC-North').teams)

        # sort division according to tiebreaker rules, highest ranked team in column 1
        z = z.xs('pct', level=1, axis=1).sort_values(list(z.index), axis=1, ascending=False)
        '''

        teams = self._teams(teams)
        df = pd.DataFrame(columns=pd.MultiIndex.from_product([teams, ['win','loss','tie', 'pct']], names=['team','outcome']))
        common_opponents = self.opponents(teams)
        divisions = set()
        stats = self._stats()

        # determine which divisions are in the specified list. If more than one then adjust the tiebreaker order
        for t in teams:
            divisions.add(self.teams_[t]['div'])

        # set rules to default values here so they appear in the correct order
        df.loc['overall'] = np.nan
        df.loc['head-to-head'] = np.nan
        if len(divisions) > 1:
            df.loc['conference'] = np.nan
            df.loc['common-games'] = np.nan
        else:
            df.loc['division'] = np.nan
            df.loc['common-games'] = np.nan
            df.loc['conference'] = np.nan
        
        df.loc['victory-strength'] = np.nan
        df.loc['schedule-strength'] = np.nan
        df.loc['conference-rank'] = np.nan
        df.loc['overall-rank'] = np.nan

        if len(divisions) > 1:
            df.loc['conference-netpoints'] = np.nan
        else:
           df.loc['common-netpoints'] = np.nan

        df.loc['overall-netpoints'] = np.nan

        (h2h,gm) = self._wlt(teams, within=teams)
        co  = self._wlt(teams, within=common_opponents)[0]

        for team in teams:
            df.loc['overall', team] = stats.loc[team,'overall'].values
            df.loc['head-to-head', team] = h2h.loc[team].drop(['scored', 'allowed']).values
            df.loc['common-games', team] = co.loc[team].drop(['scored','allowed']).values
            df.loc['conference', team] = stats.loc[team,'conference'].values
            df.loc['victory-strength', team] = stats.loc[team,'vic_stren'].values
            df.loc['schedule-strength', team] = stats.loc[team,'sch_stren'].values
            if 'division' in df.index:
                df.loc['division', team] = stats.loc[team,'division'].values

            df.loc['conference-rank', (team,'pct')] = stats.loc[team, ('misc', 'rank-conf')]
            df.loc['overall-rank', (team,'pct')] = stats.loc[team, ('misc', 'rank-overall')]
            if 'common-netpoints' in df.index:
              df.loc['common-netpoints', (team,'pct')] = co.loc[team, 'scored'] - co.loc[team, 'allowed']

            if 'conference-netpoints' in df.index:
                df.loc['conference-netpoints', (team,'pct')] = stats.loc[team, ('misc', 'conf-pts-scored')] - stats.loc[team, ('misc', 'conf-pts-allowed')]

            df.loc['overall-netpoints', (team,'pct')] = stats.loc[team, ('misc', 'pts-scored')] - stats.loc[team, ('misc', 'pts-allowed')]

            # sanity checks: we use inf to indicate that the column should be ignored
            if not gm.isin([np.nan, gm.iloc[0,1]]).all().all():
                # all teams must have played each other the same number of games or h2h is invalid
                df.loc['head-to-head', (team,'pct')] = np.inf


        # team-wide sanity checks
        if (df.loc['common-games'].drop('pct', level=1).groupby('team').sum() < 4).any():
            # if any team plays less than 4 common games, all common-team record scores are invalid
            df.loc['common-games'].loc[:, 'pct'] = np.inf

        # A note about how 'clean-sweep' analysis works
        # The sanity check below ensures that at least one team is a "clean-sweep" either perfectly winning
        # or losing against the others. If there is no "clean-sweep" (pct=[0,1]) then the rule is skipped
        #   If there is a clean-sweep loser (pct=0) that team will be dropped from contention by sorting, and the
        #   procedure restarted with the remaining teams
        #   If there is a clean-sweep winner, then the lower-ranked teams will still be dropped one-by-one
        #   as the procedure restarts i.e. a clean-sweep winner in a set of opponents is by definition
        #   a clean-sweep winner of a subset of opponents

        z = df.loc['head-to-head'].loc[:,'pct']
        if len(divisions) > 1 and len(z) > 2 and not z.isin([0, 1]).any():
            # wildcard tiebreakers with 3+ teams must be a clean sweep. If there isn't one then set to nan
            df.loc['head-to-head'].loc[:, 'pct'] = np.inf

        return df


    def tiebreaks(self, teams, fast=False):
        '''Returns a series with the results of a robust tiebreaker analysis for teams
           The series index will be in reverse elimination order, i.e. the winner (last
           eliminated team) ordered first and the loser (first eliminated team) ordered last.
           The series value corresponds to the reason each team was eliminated, relative to
           the team just before it.
        '''
        teams = list(self._teams(teams))
        r = pd.Series(name='eliminated-by')

        # shortcuts for efficiency: implement before call to _stats for speed
        if len(teams) == 0:
            return r
        elif len(teams) == 1:
            r[teams[0]] = 'winner'
            return r
        elif fast:
            # if only care about the winner, try to first discern based on overall score
            if self.stats is None:
                z = self.wlt(teams)
            else:
                z = self.stats.loc[teams,('overall','pct')].sort_values().copy()
                
            if z.iloc[0]['pct'] > z.iloc[1]['pct']:
                r[z.index[0]] = 'winner'
                return r


        stats = self._stats()
        
        def subdiv(teams):
            '''Returns a dict for the corresponding divisions of the specified teams. Dict
               keys are division codes and values are arrays of correponding team codes
            '''

            s = {}
            for elem in teams:
                d = self.teams_[elem]['div']
                if d not in s:
                    s[d] = []

                s[d].append(elem)

            return s

        def check_until_same(s):
            '''Returns the number of initial elements with the same value as the next one
            '''

            if s.isna().all() or (s == np.inf).any():
                return len(s) - 1
                
            i = 0
            for (k,v) in s.diff(-1).items():
                if v != 0:
                    return i

                i += 1

            return len(s)-1

        while len(teams) > 1:
            z = stats.loc[teams,('overall','pct')].sort_values().copy()
            t = check_until_same(z)
            if t == 0:
                r[z.index[0]] = 'overall'
                teams.remove(z.index[0])
            else:
                # check out many divisions we're about to compare
                subTeams = list(z.index[0:t+1])
                divisions = subdiv(subTeams)
    
                # may need to eliminate some teams if there are multiple divisions
                # (first team from each division only)
                if len(divisions) > 1:
                    for k,v in divisions.items():
                        if len(v) > 1:
                            s = self.tiebreaks(v)
                            
                            for t2 in reversed(s.index[1:]):
                                r[t2] = 'division-tiebreaker:' + s[t2]
                                teams.remove(t2)
                                subTeams.remove(t2)                       
                    
                z = self.tiebreakers(subTeams).xs('pct', level=1, axis=1).T
                z = z.sort_values(list(z.columns)).drop(z.columns[0], axis=1)
                z_len = len(teams)
                for rule in z.columns:
                    t = check_until_same(z[rule])
                    if t == 0:
                        r[z.index[0]] = rule
                        teams.remove(z.index[0])
                        break
    
                if z_len == len(teams):
                    raise NotImplementedError("Can't resolve tiebreaker {}: teams are essentially equal".format(len(r)))
    
        # should always be one team left
        r[teams[0]] = 'winner'
    
        # return series in reverse index order (best team first)
        return r.reindex(r.index[::-1])

    @staticmethod
    def result(a, b):
        
        if a is np.nan or b is np.nan:
            return ''
        elif a > b:
            return 'win'
        elif a < b:
            return 'loss'

        return 'tie'


    def _teams(self, teams):
        ''' Transforms teams into an array of actual team codes, or None
        '''

        if teams is None:
            return None

        if type(teams) in [list, set, range]:
            # already a list-like
            return teams

        if type(teams) in [pd.Series, pd.Index]:
            # a list-like that needs to be cast
            return list(teams)

        # below here assume a scalar

        if teams in self.divs_:
            # division code
            return self.divs_[teams]

        if teams in self.confs_:
            # conference code
            return self.confs_[teams]

        # single team code
        return [teams]

    @staticmethod
    def safeInt(i):
        '''Attempt to convert strings to integers without raising exceptions
        '''
        try:
            i = int(i)
        except ValueError:
            pass

        return i

    @staticmethod
    def to_seconds(s):
        '''Convert "mm:ss" to seconds
        '''

        (mins,secs) = map(lambda x: int(x), s.split(':'))
        return mins*60 + secs

    @staticmethod
    def to_int_list(s, sep='-'):
        '''Convert hyphenated stats to a list of integers.  "3-8-35" -> [3, 8, 35]
        '''

        return list(map(lambda x: int(x), s.split(sep)))

    def scenarios(self, weeks, teams):
        '''Iterate over all possible game outcomes
           Returns a generator that produces all possible combinations of winning
           teams in the specified weeks. Results are a dictionary that can be
           passed to set
        '''

        # subroutine based on this:
        # https://stackoverflow.com/questions/1208118/using-numpy-to-build-an-array-of-all-combinations-of-two-arrays/1235363#1235363
        def outcomes(arrays, out=None):
            arrays = [np.asarray(x) for x in arrays]
            n = np.prod([x.size for x in arrays])
            if out is None:
                out = np.zeros([n, len(arrays)], dtype='int')

            z = int(n / arrays[0].size)
            out[:,0] = np.repeat(arrays[0], z)
            if arrays[1:]:
                outcomes(arrays[1:], out=out[0:z, 1:])
                for j in range(1, arrays[0].size):
                    out[j*z:(j+1)*z, 1:] = out[0:z, 1:]

            return out

        sch = self.schedule(teams, weeks, by='game')
        for row in outcomes([(1,0,-1)] * len(sch)):
            sch['hscore'] = row
            d = {k:{} for k in sch.index.get_level_values(0)}
            for k,elem in sch.iterrows():
                d[k[0]][k[1]] = elem['hscore']

            yield d


if __name__ == '__main__':
    config = docopt(__doc__)
    # print(config)
    # sys.exit(0)

    if config['update']:
        logging.basicConfig(level=logging.INFO)

        NFL().update(config['FILE'], week=int(config['WEEK']))

    if config['team']:
        nfl = NFL()
        nfl.load(config['--file'])
        team = nfl(config['TEAM'])
        print(team)
        if type(team) is NFLTeam:
            print(team.schedule)

    if config['tiebreakers']:
        nfl = NFL()
        print(nfl.load(config['--file']).tiebreakers(config['TEAMS']))

class NFLScoreboard():
    week = None
    year = None
    scoreboard = None

    def __init__(self, year, week, scoreboard):
        self.year = year
        self.week = week
        self.scoreboard = scoreboard

    def __repr__(self):
        if type(self.scoreboard) is pd.DataFrame:
            return self.scoreboard.__repr__()

        return ''

    def _repr_html_(self):
        if type(self.scoreboard) is pd.DataFrame:
            return self.scoreboard._repr_html_()

        return ''

class NFLSource():
    lasturl = None

    def boxscore(self, nfl, code, week):
        '''Return box score as a data frame
        '''

        raise NotImplementedError('{} does not implement box scores'.format(self.__class__.__name__))

    def scoreboard(self, nfl):

        raise NotImplementedError('{} does not implement scoreboard'.format(self.__class__.__name__))


class NFLSourceProFootballRef(NFLSource):

    source = 'https://www.pro-football-reference.com/years/{}/games.htm'

    def games(self, nfl):

        url = self.source.format(nfl.year)

        try:
            d = PyQuery(url=url)('#all_games #games tbody > tr')
        except urllib.error.HTTPError as err:
            logging.error('Bad URL: {}'.format(url))
            raise

        def rowval(elem, id, tag='td'):
            return PyQuery(elem)('{}[data-stat="{}"]'.format(tag, id)).text()

        self.lasturl = url
        for elem in d:
            # NB: the tag attributes may be different prior to season start. See prebuild()
            week = NFL.safeInt(rowval(elem, "week_num", "th"))
            if type(week) is int:
                aname  = rowval(elem, "winner")
                hname  = rowval(elem, "loser")
                ascore = NFL.safeInt(rowval(elem, "pts_win"))
                hscore = NFL.safeInt(rowval(elem, "pts_lose"))
                game_date = rowval(elem, "game_date")
                game_time = rowval(elem, "gametime")
                game_date = datetime.strptime(game_date+game_time, "%Y-%m-%d%I:%M%p")

                if rowval(elem, "game_location") != '@':
                    (aname,hname) = (hname,aname)
                    (ascore,hscore) = (hscore,ascore)

                yield {
                    'week': week,
                    'date': game_date,
                    'ateam': aname,
                    'hteam': hname,
                    'ascore': ascore,
                    'hscore': hscore,
                }

    def boxscore(self, nfl, code, week):
        '''Return a dataframe of games stats for the specified week. None is returned for bye weeks
        '''

        game = nfl.game(code, week)
        if game:
            ht = game['ht']
            alt = nfl.teams_[ht]['alt'] or ht
            code = '{}0{}'.format(datetime.strftime(game['ts'], '%Y%m%d'), alt.lower())

            setup = {
                'head': ['week', 'opp', 'date','id'],
                'points': ['q1', 'q2', 'q3', 'q4', 'ot', 'final'],
                'game': ['yards', '1st_downs', 'turnovers', 'time_of_poss', 'secs_of_poss'],
                'rushing': ['count', 'yds', 'tds'],
                'passing': ['comp', 'att', 'yds', 'yds_net', 'tds', 'int'],
                'fumbles': ['count', 'lost'],
                'sacks': ['count', 'yds_lost'],
                '3d_conv': ['count', 'of'],
                '4d_conv': ['count', 'of'],
                'penalties': ['count', 'yds'],
            }

            idx = pd.MultiIndex.from_arrays([
                list(np.concatenate([[k]*len(v) for k,v in setup.items()])),
                list(np.concatenate(list(setup.values())))
                ], names=['cat', 'key'])

            df = pd.DataFrame(index=idx, columns=[game['at'], game['ht']])
            df.loc[('head','week'), :] = game['wk']
            df.loc[('head','opp'), :] = [game['ht'], game['at']]
            df.loc[('head','date'), :] = datetime.strftime(game['ts'], '%Y-%m-%d')
            df.loc[('head','id'), :] = code

            if not game['p']:
                return df.dropna()
            
            url = 'https://www.pro-football-reference.com/boxscores/{}.htm'.format(code)
            d = PyQuery(url=url)
            self.lasturl = url

            t = PyQuery(d)('table.linescore')
            quarters = [PyQuery(elem).text().lower() for elem in PyQuery(t)('thead th')][2:]
            ateam    = [PyQuery(elem).text().lower() for elem in PyQuery(t)('tr:nth-child(1) td')][2:]
            hteam    = [PyQuery(elem).text().lower() for elem in PyQuery(t)('tr:nth-child(2) td')][2:]

            quarters[0:4] = ['q'+x for x in quarters[0:4]]
            for i in zip(quarters, ateam, hteam):
                df.loc[('points', i[0]), :] = [int(i[1]), int(i[2])]

            # #team_stats table is commented out, so we have to dig for it
            html = PyQuery(d)('#all_team_stats').html().replace('<!--', '').replace('-->', '')
            team_stats = PyQuery(html)('#team_stats')

            for tr in PyQuery(team_stats)('tr'):
                stat = PyQuery(tr)('th').eq(0).text().lower()
                at   = PyQuery(tr)('td').eq(0).text().lower()
                ht   = PyQuery(tr)('td').eq(1).text().lower()

                if stat == 'first downs':
                    df.loc[('game', '1st_downs'), :] = [int(at), int(ht)]
                elif stat == 'total yards':
                    df.loc[('game', 'yards'), :] = [int(at), int(ht)]
                elif stat == 'turnovers':
                    df.loc[('game', 'turnovers'), :] = [int(at), int(ht)]
                elif stat == 'penalties-yards':
                    at = NFL.to_int_list(at)
                    ht = NFL.to_int_list(ht)
                    df.loc[('penalties','count'), :] = [at[0], ht[0]]
                    df.loc[('penalties','yds'), :]   = [at[1], ht[1]]
                elif stat == 'sacked-yards':
                    at = NFL.to_int_list(at)
                    ht = NFL.to_int_list(ht)
                    df.loc[('sacks','count'), :] = [at[0], ht[0]]
                    df.loc[('sacks','yds_lost'), :]   = [at[1], ht[1]]
                elif stat == 'fumbles-lost':
                    at = NFL.to_int_list(at)
                    ht = NFL.to_int_list(ht)
                    df.loc[('fumbles','count'), :] = [at[0], ht[0]]
                    df.loc[('fumbles','lost'), :]   = [at[1], ht[1]]
                elif stat == 'rush-yds-tds':
                    at = NFL.to_int_list(at)
                    ht = NFL.to_int_list(ht)
                    df.loc[('rushing','count'), :] = [at[0], ht[0]]
                    df.loc[('rushing','yds'), :]   = [at[1], ht[1]]
                    df.loc[('rushing','tds'), :]   = [at[2], ht[2]]
                elif stat == 'cmp-att-yd-td-int':
                    at = NFL.to_int_list(at)
                    ht = NFL.to_int_list(ht)
                    df.loc[('passing','comp'), :] = [at[0], ht[0]]
                    df.loc[('passing','att'), :]   = [at[1], ht[1]]
                    df.loc[('passing','yds'), :]   = [at[2], ht[2]]
                    df.loc[('passing','tds'), :]   = [at[3], ht[3]]
                    df.loc[('passing','int'), :]   = [at[4], ht[4]]
                elif stat == 'net pass yards':
                    df.loc[('passing', 'yds_net'), :] = [int(at), int(ht)]
                elif stat == 'time of possession':
                    df.loc[('game', 'time_of_poss'), :] = [at, ht]
                    df.loc[('game', 'secs_of_poss'), :] = [NFL.to_seconds(at), NFL.to_seconds(ht)]
                elif stat in ['third down conv.', 'fourth down conv.']:
                    key = '3d_conv' if stat == 'third down conv.' else '4d_conv'
                    at = NFL.to_int_list(at)
                    ht = NFL.to_int_list(ht)
                    df.loc[(key,'count'), :] = [at[0], ht[0]]
                    df.loc[(key,'of'), :]   = [at[1], ht[1]]

            return df



class NFLSourceESPN(NFLSource):

    source = 'https://site.api.espn.com/apis/site/v2/sports/football/nfl/scoreboard?dates={:04d}{:02d}'

    def __init__(self):
        self.zone = tz.gettz('America/New_York')

    def games(self, nfl):

        season_type = 2
        (start,end) = self.season_dates(nfl.year, season_type)
        first = start.floor(freq='D').replace(day=1)
        last  = end.floor(freq='D').replace(day=1)

        d = first
        while d <= last:
            self.lasturl = self.source.format(d.year, d.month)
            result = requests.get(self.lasturl).json()
            for elem in result['events']:
                if elem['season']['type'] == season_type:

                    (hteam,ateam) = elem['competitions'][0]['competitors']
                    if hteam['homeAway'] != 'home':
                        (hteam,ateam) = (ateam,hteam)

                    if elem['competitions'][0]['status']['type']['completed']:
                        ascore = NFL.safeInt(ateam['score'])
                        hscore = NFL.safeInt(hteam['score'])
                    else:
                        ascore = hscore = None

                    yield {
                        'week': elem['week']['number'],
                        'date': self.to_datetime(elem['date']),
                        'ateam': ateam['team']['displayName'],
                        'hteam': hteam['team']['displayName'],
                        'ascore': ascore,
                        'hscore': hscore,
                        'id': elem['id']
                    }

            d += pd.DateOffset(months=1)

    def boxscore(self, nfl, code, week):

        def team_stats(t):

            z = t['team']
            z['stats'] = {}
            for elem in t['statistics']:
                z['stats'][elem['name']] = elem['displayValue']

            return z

        game = nfl.game(code, week)
        if game:
            url = 'https://site.api.espn.com/apis/site/v2/sports/football/nfl/summary?event={}'
            self.lasturl = url.format(game['id'])
            result = requests.get(self.lasturl).json()

            setup = {
                'head': ['week', 'opp', 'date', 'event'],
                'points': ['q1', 'q2', 'q3', 'q4', 'ot', 'final'],
                'game': ['drives', 'yards', '1st_downs', 'turnovers', 'fumbles_lost', 'time_of_poss', 'secs_of_poss'],
                'rushing': ['count', 'yds', 'tds'],
                'passing': ['comp', 'att', 'yds_net', 'tds', 'int'],
                'sacks': ['count', 'yds_lost'],
                '3d_conv': ['count', 'of'],
                '4d_conv': ['count', 'of'],
                'penalties': ['count', 'yds'],
            }

            idx = pd.MultiIndex.from_arrays([
                list(np.concatenate([[k]*len(v) for k,v in setup.items()])),
                list(np.concatenate(list(setup.values())))
                ], names=['cat', 'key'])

            df = pd.DataFrame(index=idx, columns=[game['at'], game['ht']])
            df.loc[('head','week'), :] = game['wk']
            df.loc[('head','opp'), :] = [game['ht'], game['at']]
            df.loc[('head','date'), :] = datetime.strftime(game['ts'], '%Y-%m-%d')
            df.loc[('head','event'), :] = game['id']

            if not game['p']:
                return df.dropna()

            (hteam,ateam) = result['header']['competitions'][0]['competitors']
            if hteam['homeAway'] != 'home':
                (hteam,ateam) = (ateam,hteam)

            df.loc[('points','final'), :] = [NFL.safeInt(ateam['score']), NFL.safeInt(hteam['score'])]

            qtrs = list(df.index.get_loc_level('points')[1])
            for i in zip(qtrs, ateam['linescores'], hteam['linescores']):
                df.loc[('points', i[0]), :] = [NFL.safeInt(i[1]['displayValue']), NFL.safeInt(i[2]['displayValue'])]

            # unpack team statistics into something simpler
            (hstats,astats) = map(lambda x: x['statistics'], result['boxscore']['teams'])
            if result['boxscore']['teams'][0]['homeAway'] != 'home':
                (hstats,astats) = (astats,hstats)

            hstats = {elem['name']: elem['displayValue'] for elem in hstats}
            astats = {elem['name']: elem['displayValue'] for elem in astats}
            stats  = {k:[astats.get(k), hstats.get(k)] for k in hstats}

            df.loc[('game','drives'), :] = stats['totalDrives']
            df.loc[('game','yards'), :] = stats['totalYards']
            df.loc[('game','1st_downs'), :] = stats['firstDowns']
            df.loc[('game','turnovers'), :] = stats['turnovers']
            df.loc[('game', 'fumbles_lost'), :] = stats['fumblesLost']
            df.loc[('game','time_of_poss'), :] = stats['possessionTime']
            df.loc[('game','secs_of_poss'), :] = [NFL.to_seconds(stats['possessionTime'][0]), NFL.to_seconds(stats['possessionTime'][1])]

            at = NFL.to_int_list(stats['sacksYardsLost'][0])
            ht = NFL.to_int_list(stats['sacksYardsLost'][1])
            df.loc[('sacks','count'), :] = [at[0], ht[0]]
            df.loc[('sacks','yds_lost'), :]   = [at[1], ht[1]]

            at = NFL.to_int_list(stats['totalPenaltiesYards'][0])
            ht = NFL.to_int_list(stats['totalPenaltiesYards'][1])
            df.loc[('penalties','count'), :] = [at[0], ht[0]]
            df.loc[('penalties','yds'), :]   = [at[1], ht[1]]

            df.loc[('rushing','count'), :] = stats['rushingAttempts']
            df.loc[('rushing','yds'), :] = stats['rushingYards']
            df.loc[('rushing','tds'), :] = 0

            at = NFL.to_int_list(stats['completionAttempts'][0], '/')
            ht = NFL.to_int_list(stats['completionAttempts'][1], '/')

            df.loc[('passing','comp'), :] = [at[0], ht[0]]
            df.loc[('passing','att'), :] = [at[1], ht[1]]
            df.loc[('passing','yds_net'), :] = stats['netPassingYards']
            df.loc[('passing','int'), :] = stats['interceptions']
            df.loc[('passing','tds'), :] = 0

            at = NFL.to_int_list(stats['thirdDownEff'][0])
            ht = NFL.to_int_list(stats['thirdDownEff'][1])
            df.loc[('3d_conv', 'count'), :] = [at[0], ht[0]]
            df.loc[('3d_conv', 'of'), :] = [at[1], ht[1]]

            at = NFL.to_int_list(stats['fourthDownEff'][0])
            ht = NFL.to_int_list(stats['fourthDownEff'][1])
            df.loc[('4d_conv', 'count'), :] = [at[0], ht[0]]
            df.loc[('4d_conv', 'of'), :] = [at[1], ht[1]]

            # iterate scoringPlays to count touchdowns
            for score in result['scoringPlays']:
                key = game['ht'] if score['team']['uid'] == hteam['uid'] else game['at']
                if score['type']['id'] == '68':
                    df.loc[('rushing','tds'), key] += 1
                elif score['type']['id'] == '67':
                    df.loc[('passing','tds'), key] += 1

            return df

    def scoreboard(self, nfl):
        '''Return current scoreboard
        '''

        df = pd.DataFrame(columns=['hteam','ateam','hscore','ascore','period','clock','status','down','fpos','broadcast'])
        now = datetime.now()
        # url = 'https://site.api.espn.com/apis/site/v2/sports/football/nfl/scoreboard?dates={:04d}{:02d}{:02d}'
        # self.lasturl = url.format(now.year, now.month, now.day)
        # result = requests.get(self.lasturl).json()
        self.lasturl = 'https://site.api.espn.com/apis/site/v2/sports/football/nfl/scoreboard'

        def rget(obj, *argv):
            '''Safely retrieves values from a nested set of dicts
            '''
            if len(argv) == 0:
                return obj

            if type(obj) is dict:
                return rget(obj.get(argv[0]), *argv[1:])

            return None

        result = requests.get(self.lasturl).json()
        for elem in result['events']:
            game = elem['competitions'][0]
            (hteam,ateam) = game['competitors']
            if hteam['homeAway'] != 'home':
                (hteam,ateam) = (ateam,hteam)

            at = len(df)
            df.loc[at, ['hteam','ateam']] = [nfl.teamid(hteam['team']['displayName']), nfl.teamid(ateam['team']['displayName'])]
            df.loc[at, ['hscore','ascore']] = [NFL.safeInt(hteam['score']), NFL.safeInt(ateam['score'])]
            df.loc[at, ['broadcast','down','fpos','period','clock']] = [game['broadcast'], '','','','']
            df.loc[at, 'status'] = game['status']['type']['detail'] # default
            status = game['status']['type']['state']
            if status == 'in':
                df.loc[at, ['period','clock']] = [game['status']['period'], game['status']['displayClock']]
                if type(game.get('situation')) is dict:
                    sit = game['situation']
                    if sit.get('possession'):
                        pos = sit['possession']
                        pos = df.loc[at, 'hteam'] if pos == hteam['id'] else df.loc[at, 'ateam']
                        df.loc[at, ['status','down','fpos']] = [pos, sit.get('shortDownDistanceText',''), sit.get('possessionText','')]
                    elif game['status']['type']['name'] == 'STATUS_HALFTIME':
                        df.loc[at, 'status'] = game['status']['type']['shortDetail']
                    elif rget(sit, 'lastPlay', 'drive', 'result'):
                        # between possessions: try to report result of last play
                        pos = sit['lastPlay']['team']['id']
                        pos = df.loc[at, 'hteam'] if pos == hteam['id'] else df.loc[at, 'ateam']
                        df.loc[at, 'status'] = '{} {}'.format(sit['lastPlay']['drive']['result'], pos)
                    else:
                        df.loc[at, 'status'] = rget(sit, 'lastPlay', 'type', 'text') or 'na'

            elif status == 'pre':
                df.loc[at, ['hscore','ascore']] = np.nan
                gametime = pd.to_datetime(game['date']).astimezone(self.zone).replace(tzinfo=None)
                df.loc[at, 'status'] = '{}/{} {:02d}:{:02d}'.format(gametime.month, gametime.day, gametime.hour, gametime.minute)

        return NFLScoreboard(result['season']['year'], result['week']['number'], df)


    def to_datetime(self, date):
        '''Returns string converted to a zoneless datetime in the current time zone
        '''
        return pd.to_datetime(date).to_pydatetime().astimezone(self.zone).replace(tzinfo=None)

    def season_dates(self, year, type=2):
        '''Returns span for a season (regular season = 2) as pandas datetimes
        '''

        url = 'https://sports.core.api.espn.com/v2/sports/football/leagues/nfl/seasons/{}/types/{}'.format(year, type)
        result = requests.get(url).json()
        return (pd.to_datetime(result['startDate']), pd.to_datetime(result['endDate']))
